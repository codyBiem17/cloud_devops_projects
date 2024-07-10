import openpyxl as pyxl

inv_file = pyxl.load_workbook('inventory_file_nana.xlsx')
product_list = inv_file['Sheet1']
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_no = product_list.cell(product_row, 1).value
    inv_price = product_list.cell(product_row, 5)

    # calculation of num of prod per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name] + 1
        # you can also use the 'get' function instead of [] to get value

        # assign new value to num of products
        products_per_supplier[supplier_name] = current_num_products
    else:
        products_per_supplier[supplier_name] = 1

    # calculation of total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name) + inventory * price
        total_value_per_supplier[supplier_name] = current_total_value
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculation of products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_no)] = int(inventory)

    # create a new file with inventory_price column from existing inventory file
    inv_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("new_inv_file.xlsx")


